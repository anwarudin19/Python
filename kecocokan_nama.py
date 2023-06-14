from flask import Flask, request
from jellyfish import jaro_winkler_similarity
from jellyfish import jaro_similarity

from flask import Flask, jsonify, request
from flask_cors import CORS, cross_origin
import openpyxl
import logging

import json

# Konfigurasi logger

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Buat handler untuk menyimpan log ke file
file_handler = logging.FileHandler('log.txt')
file_handler.setLevel(logging.DEBUG)

# Atur format log
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Tambahkan handler ke logger
logger.addHandler(file_handler)

# Inisialisasi aplikasi Flask
app = Flask(__name__)
def calculate_similarity(nama1, nama2, gender1, gender2):
    # similarity = jaro_winkler_similarity(nama1, nama2)
    similarity = jaro_similarity(nama1, nama2)
    similarity_percent = similarity * 100
    gender1 *=1.1
    gender2 *=0.9
    return similarity_percent
CORS(app)
app.debug = True
CORS(app, resources={r"/*": {"origins": "*"}})
@cross_origin()
@app.route('/kecocokan_nama', methods=['POST'])

# Route untuk endpoint API
def kecocokan():
    data = request.get_json()

    if 'nama1' not in data or 'nama2' not in data:
        return 'Field nama1 atau nama2 tidak ditemukan dalam data', 400

    nama1 = data['nama1']
    nama2 = data['nama2']
    gender1 = 1.1
    gender2 = 0.9
    similarity = calculate_similarity(nama1, nama2, gender1, gender2)
    # Log pesan
    logger.info('Memproses kecocokan untuk nama {} dan {} {} {}'.format(nama1, nama2,gender1,gender2))

    # Menghitung similarity antara kedua nama menggunakan metode Jaro
    # similarity = jaro_similarity(nama1, nama2)
    response = {'similarity': f'{similarity:.2f}%'}
    
    with open('D:/data/Github/python/Belajar/data.json', 'r') as file:
        data_json = file.read()
    

    data_json = json.loads(data_json)
    json_result = {
        "nama1" : data["nama1"],
        "nama2" : data["nama2"],    
        "similarity" : f"{similarity:.2f}%",
    }
    data_json["data"].append(json_result)
    
    new_json = json.dumps(data_json)
    with open('D:/data/Github/python/Belajar/data.json', 'w') as file:
        file.write(new_json)
    # Tulis hasil ke dalam file Excel
    # tulis_hasil_ke_excel(nama1, nama2, similarity)
    return jsonify(response)
def tulis_hasil_ke_excel(nama1, nama2, similarity):
    # Buka file Excel atau buat file baru jika belum ada
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Tulis header kolom
    sheet["A1"] = "Nama Pertama"
    sheet["B1"] = "Nama Kedua"
    sheet["C1"] = "Persentase Kecocokan"

    # Tulis data hasil kecocokan nama
    row = sheet.max_row + 1
    print(row)
    sheet.cell(row=row, column=1, value=nama1)
    sheet.cell(row=row, column=2, value=nama2)
    sheet.cell(row=row, column=3, value=similarity)

    # Simpan file Excel
    workbook.save("hasil_kecocokan.csv")
    workbook.close()
    
@app.route('/kecocokan_nama', methods=['GET'])  
def get_kecocokan_nama():
    with open('D:/data/Github/python/Belajar/data.json', 'r') as file:
        data_json = file.read()
    # print(data_json)
    data_json = json.loads(data_json)
    data_json["data"] = list(data_json["data"])[::-1]
    return data_json


if __name__ == '__main__':
    app.run()
